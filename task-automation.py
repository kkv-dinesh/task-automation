import os
import openpyxl
import win32com.client as win32
from PyPDF2 import PdfMerger
from PIL import Image

def send_email_with_attachment(attachment_path, recipient_name, recipient_email, cc_email):
    # Get the Outlook application object
    outlook = win32.Dispatch('outlook.application')

    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.CC = cc_email

    # Set the email subject
    mail.Subject = f"Financial Data: {attachment_path}"

    # Set the email text
    mail.Body = f"Dear {recipient_name},\n\nPlease find the attached financial data for {attachment_path}.\n\nBest regards,\nDinesh"

    # Add the attachment
    mail.Attachments.Add(attachment_path)

    # Send the email
    mail.Send()


def extract_pages_as_text(pdf_file, start_page, end_page, output_file):
    with open(pdf_file, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        # Validate start and end page numbers
        if start_page < 1 or end_page > num_pages or start_page > end_page:
            print("Invalid page range.")
            return

        extracted_text = ""
        for page_number in range(start_page - 1, end_page):
            page = pdf_reader.pages[page_number]
            extracted_text += page.extract_text()

        # Save extracted text to a text file
        with open(output_file, 'w') as output:
            output.write(extracted_text)
        
        print("Text extracted and saved to", output_file)

def merge_pdfs(cover_folder, draft_folder, output_folder):
    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.mkdir(output_folder)

    # Get a list of PDFs in the cover and draft folders
    cover_pdfs = [f for f in os.listdir(cover_folder) if f.endswith('.pdf')]
    draft_pdfs = [f for f in os.listdir(draft_folder) if f.endswith('.pdf')]

    # Iterate through the cover PDFs
    for pdf in cover_pdfs:
        # Get the numeric key of the PDF
        key = pdf.split('.')[0]

        # Find the matching draft PDF with the same key
        matching_draft = [f for f in draft_pdfs if key in f]
        
        if len(matching_draft) == 0:
            print(f"No matching draft PDF found for {pdf}. Skipping...")
            continue
        
        matching_draft = matching_draft[0]

        # Create a PdfMerger object
        merger = PdfMerger()

        # Add the cover and draft PDFs to the merger
        merger.append(os.path.join(cover_folder, pdf))
        merger.append(os.path.join(draft_folder, matching_draft))

        # Write the merged PDF to the output folder
        output_path = os.path.join(output_folder, matching_draft)
        merger.write(output_path)

        # Close the merger object
        merger.close()

    print(f'Merged {len(cover_pdfs)} PDFs to {output_folder}')

def perform_image_processing():
    # Get the input image path from the user
    input_path = "F:\img\pexels-sebastiaan-stam-1097456.jpg"

    # Get the output image path from the user
    output_path ="F:\img\output.jpg"
    # Open the input image
    input_image = Image.open(input_path)

    # Get the desired width and height for resizing from the user
    new_width = int(input("Enter the desired width for resizing: "))
    new_height = int(input("Enter the desired height for resizing: "))

    # Resize the image
    resized_image = input_image.resize((new_width, new_height))

    # Get the rotation angle from the user
    angle = float(input("Enter the rotation angle (in degrees): "))

    # Rotate the image
    rotated_image = resized_image.rotate(angle)

    # Get the crop coordinates from the user
    left = int(input("Enter the left coordinate for cropping: "))
    upper = int(input("Enter the upper coordinate for cropping: "))
    right = int(input("Enter the right coordinate for cropping: "))
    lower = int(input("Enter the lower coordinate for cropping: "))

    # Crop the image
    cropped_image = rotated_image.crop((left, upper, right, lower))

    # Save the final processed image
    cropped_image.save(output_path)

    # Display a success message
    print("Image processing complete. The processed image is saved at", output_path)

# Menu-driven program
while True:
    print("---- Menu ----")
    print("1. Email Automation")
    print("2. PDF Text Extraction")
    print("3. PDF Merging")
    print("4. Image Processing")
    print("5. Exit")

    choice = input("Enter your choice (1-5): ")

    if choice == '1':
        # Email Automation Task
        cwd = "D:\Project\excel"
        workbook = openpyxl.load_workbook(os.path.join(cwd, "Financial_Data.xlsx"))
        sheet = workbook["Email_List"]

        for i in range(2, sheet.max_row + 1):
            attachment = sheet.cell(row=i, column=1).value
            attachment_path = os.path.join(cwd, attachment)
            if not os.path.exists(attachment_path):
                print(f"Attachment {attachment} does not exist")
                continue

            recipient_name = sheet.cell(row=i, column=2).value
            recipient_email = sheet.cell(row=i, column=3).value
            cc_email = sheet.cell(row=i, column=4).value

            send_email_with_attachment(attachment_path, recipient_name, recipient_email, cc_email)

            print("Mail Sent !!!")
        workbook.close()

    elif choice == '2':
        # PDF Text Extraction Task
        pdf_file_path = "D:\Project\computer_networks_-_a_tanenbaum_-_5th_edition.pdf"
        start_page_number = int(input("Enter the starting page number: "))
        end_page_number = int(input("Enter the ending page number: "))
        output_file_path = "D:\\Project\\output.txt"

        extract_pages_as_text(pdf_file_path, start_page_number, end_page_number, output_file_path)

    elif choice == '3':
        # PDF Merging Task
        cover_folder = r'D:\sem-4\cf lab\Pdf\cover'
        draft_folder = r'D:\sem-4\cf lab\Pdf\draft'
        output_folder = r'D:\Project\prj'

        merge_pdfs(cover_folder, draft_folder, output_folder)

    elif choice == '4':
        # Image Processing Task
        perform_image_processing()

    elif choice == '5':
        # Exit the program
        break

    else:
        print("Invalid choice. Please enter a valid option (1-5).")

print("Program exited.")
